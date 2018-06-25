# -*- coding: utf-8 -*-
from __future__ import unicode_literals

import random
from io import BytesIO

from django.contrib.auth import authenticate, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect
from django.shortcuts import render, redirect
from django.contrib.auth import login as django_login
from django.contrib.auth.models import User as Django_User

from users.models import Profile


def send_email(to, subject, body):
    import smtplib

    # IN views.py you have to provice the credentials

    gmail_user = "magicampus1@gmail.com"  # email from which account you want to send emails
    gmail_pwd = "Admin123!"  # password of above email
    FROM = "magicampus1@gmail.com"
    # This is the list of users who will recieve the email
    # If I add your email address here you will start getting emails
    recipient = to  # list of email recievers
    TO = recipient if type(recipient) is list else [recipient]
    SUBJECT = subject
    TEXT = body

    # Prepare actual message
    message = """From: %s\nTo: %s\nSubject: %s\n\n%s
    """ % (FROM, ", ".join(TO), SUBJECT, TEXT)
    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.ehlo()
        server.starttls()
        server.login(gmail_user, gmail_pwd)
        server.sendmail(FROM, TO, message)
        server.close()
        print('successfully sent the mail')
        return True
    except:
        print("failed to send mail")
        return False


def design_message(name, email, number, message):
    formatted_message = "Hello !\nA New Message from Magicapus has been recieved\n"
    formatted_message += "Submitted by: " + name + "\n"
    formatted_message += "Contact Information: \n  Phone Number: " + number + "\n  Email: " + email + "\n"
    formatted_message += "Message :\n  " + message + "\n Thank You!"
    return formatted_message


def design_activation_email(request, hash):
    formatted_message = "Click here to activate your Account: http://" + str(
        request.META['HTTP_HOST']) + "/activate?hash=" + str(hash)
    return formatted_message


# Create your views here.
def index(request):
    try:
        if request.user.is_authenticated():
            return HttpResponseRedirect('/home')
    except:
        pass
    error = False
    email_sent = False
    if request.method == 'POST':
        name = request.POST.get('name') if "name" in request.POST else None
        sender = request.POST.get('sender') if "sender" in request.POST else None
        number = request.POST.get('number') if "number" in request.POST else None
        message = request.POST.get('message') if "message" in request.POST else None

        if name and sender and number and message:
            formatted_message = design_message(name, sender, number, message)
            print(formatted_message)
            email_sent = send_email(["shahraizali10@yahoo.com"],"Email From Magicampus", formatted_message)
            if not email_sent:
                error = True

    return render(request, 'home/index.html', {
        "error": error,
        "email_sent": email_sent
    })


def subscribe(request):
    return render(request, "home/subscribe.html", {})


def donate(request):
    return render(request, "home/donate.html", {})

def custom_auth(username, password):
    try:
        user = authenticate(username=username, password=password)
        if user is not None:
            return user
        else:
            django_user = Django_User.objects.get(username=username)
            django_user.backend = "django.contrib.auth.backends.ModelBackend"
            if django_user.is_active:
                profile = Profile.objects.get(id=django_user.id)
                if profile.password == password:
                    return django_user
    except:
        return None

def login(request):
    try:
        if request.user.is_authenticated():
            return HttpResponseRedirect('/home')
    except:
        pass
    message = {}
    if request.method == 'POST':
        username = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        if username and password:
            try:
                user = custom_auth(username=username, password=password)
                if user is not None:
                    if user.is_active:
                        django_login(request, user, backend="django.contrib.auth.backends.ModelBackend")
                        return redirect('home')
                else:
                    try:
                        django_user = Django_User.objects.get(username=username)
                        if django_user.is_active:
                            profile = Profile.objects.get(id=django_user.id)
                            if profile.password == password:
                                django_login(request, django_user)
                                return redirect('home')
                    except:
                        pass
                    message["title"] = "Error"
                    message["type"] = "danger"
                    message["msg"] = "Login Failed"
            except:

                message["title"] = "Error"
                message["type"] = "danger"
                message["msg"] = "Authentication error"
    return render(request, "login.html", {"message": message})


def signup(request):
    message = {}
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()

        if password and confirm_password:
            if password == confirm_password:
                # Creating login profile for incoming user/organization
                try:
                    django_user = Django_User.objects.create(username=str(email).lower(),
                                                             first_name=first_name, last_name=last_name,
                                                             email=email,
                                                             is_active=False
                                                             )
                    django_user.set_password(password)
                    django_user.save()
                    hash = random.getrandbits(128)
                    Profile.objects.create(
                        user=django_user,
                        hash=hash,
                        is_subscriber=False,
                        password=password
                    )
                    formated_message = design_activation_email(request, hash)
                    send_email([django_user.email], "activation Mail", formated_message)
                    message["title"] = "Success"
                    message["type"] = "warning"
                    message[
                        "msg"] = "An activation Email has been Sent to your Email Address. Please Confirm your Email before Logging In"
                    message["signup"] = "success"
                except :
                    message["title"] = "Error"
                    message["type"] = "danger"
                    message["msg"] = "Email Already Exist"
            else:
                message["title"] = "Error"
                message["type"] = "danger"
                message["msg"] = "Password Didn't match"

    return render(request, "signup.html", {
        "message": message
    })


@login_required
def signout(request):
    logout(request)
    request.session.flush()
    return redirect('index')


def activate(request):
    hash = request.GET["hash"]
    try:
        user = Profile.objects.get(hash=hash)
        if user:
            user.user.is_active = True
            user.user.save()
            try:
                django_login(request, user)
                return redirect('index')
            except:
                pass
    except Profile.DoesNotExist:
        user = ""
    return redirect("index")


@login_required
def home(request):
    return render(request, "home/home.html", {})


@login_required
def bulk_users(request):
    if request.user.is_staff:
        from openpyxl import load_workbook
        reports = []
        if request.method == "POST" and request.FILES['file_name']:
            file_in_memory = request.FILES['file_name'].read()
            workbook = load_workbook(filename=BytesIO(file_in_memory))

            try:

                worksheet = workbook.get_sheet_by_name('users')

                for i in range(worksheet._current_row):
                    report = {}
                    email = worksheet.cell(i + 2, 1).value
                    password = worksheet.cell(i + 2, 2).value
                    first_name = worksheet.cell(i + 2, 3).value
                    last_name = worksheet.cell(i + 2, 4).value
                    if email != None:
                        try:
                            django_user = Django_User.objects.create(username=str(email).lower(),
                                                                     first_name=first_name, last_name=last_name,
                                                                     email=email,
                                                                     is_active=False
                                                                     )
                            django_user.set_password(password)
                            django_user.save()
                            hash = random.getrandbits(128)
                            Profile.objects.create(
                                user=django_user,
                                hash=hash,
                                is_subscriber=False,
                                password=password
                            )
                            formated_message = design_activation_email(request, hash)
                            send_email([django_user.email], "activation Mail", formated_message)
                            report["type"] = "success"
                            report["message"] = "User '"+ str(email)+ "' Added successfully  at Ln:"+str(i+2)
                            reports.append(report)
                        except:
                            report["type"] = "danger"
                            report["message"] = "Error occured while adding User '"+ str(email)+ "'! at Ln:"+str(i+2)
                            reports.append(report)
            except:
                pass
        return render(request, "bulk_users.html", {
            "reports": reports
        })
    else:
        return redirect("home")